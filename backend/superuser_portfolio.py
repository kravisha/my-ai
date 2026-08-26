"""The operator's own portfolio, owned (TASK_QUEUE TQ-46, addendum 44 §4, §10,
§16, §21.4, §21.6).

Source: `docs/specs/TQ-46_superuser_ownership_domain.md`; owner direction
2026-08-26 on who the Superuser is (§11 Q4).

## What this replaces

`app/tools/portfolio.py::retrieve_portfolio` took a permission manager, a
preference store and an audit log - and **no owner**. It read
`data/portfolio.xlsx` and returned whatever was in it. That was §16's "legacy
single-portfolio design", and it was the last place in this codebase where
portfolio data was reached without anybody asking whose it was.

TQ-44 built the machinery that makes an answer possible; nothing had used it.
This is the thing that uses it.

## Whose portfolio this is, and why the answer is not "the operator"

Decided by the owner on 2026-08-26: **the backend `username`** - the account
`/chat` resolves with `Depends(get_current_user)`, which TQ-69 §3 found already
resolved and thrown away.

The owner's framing settled the structural question rather than the string:
*"Krish is the superuser of this system… the architect… the owner… the creator.
The system serves Krish primarily. The system also has other clients… However,
they are just clients. The clients pay for service."* **Krish is not in the
clients population.** `gateway/clients.py` is the registry of external, paying
clients; `users.json` is the system's own principal.

## One person, two doors - and that is the design, not an accident

`GATEWAY_SUPER_USER` and the backend username both read `krish`. **That is
intentional** - owner direction, 2026-08-26 - and it was worth asking about,
because an identical string arrived at from two independent places is exactly
what a coincidence looks like from the inside.

It is not one. There is one operator identity, reachable through two doors: the
Gateway for the outside, `/chat` and the console for the system's own surfaces.
So the job is **not** to pick a winner between two ids. It is to make the sameness
something the system *holds* rather than something a deployment happens to get
right.

Two mechanisms do that, and neither is a comment saying "keep these in sync":

- **`gateway/auth.subject_for`** records the *configured* operator name as the
  session subject, never the string that was typed. So the owner's `superuser`
  pseudo-login and their own name produce one subject, and it is the same name
  the backend knows them by.
- **`normalise` has one implementation** (TQ-69), so the two sides cannot
  disagree over case either.

What stays true regardless is that **only this process asserts `SUPERUSER`
ownership.** `/chat` is here; TQ-47's Superuser Portfolio tab is *"a dedicated
tab in the console"*, which is this process's own surface, and the Gateway
already proxies it as the studio. The Gateway gets no superuser-portfolio tool -
not because its identity is untrustworthy, but because the operator reaching
their own portfolio through the *client* door is the ambiguity §3 and §4.4 warn
about, arriving one layer lower.

The failure this arrangement exists to prevent has not gone away, and it is worth
naming precisely: a mismatch between the id the migration stamped and the id a
caller asserts is **not refused**. `resolve()` answering "not yours" is
indistinguishable from "no such portfolio" by design (addendum 44 §9.3), so the
operator would be shown an **empty portfolio and a working-looking system**.
`test_the_operator_is_never_silently_shown_an_empty_portfolio` is the regression
with that shape.

## Two portfolios, one person, named apart

The operator can end up with two: the holdings they stated through the Gateway as
a client (§96), and this one. That is not a breach - `_owned_by` compares
`(owner_type, owner_id)` as a pair, so one name in two domains is two owners, and
each refuses the other. It is an *ambiguity*, which is its own failure: somebody
being shown the wrong money.

§96 already said naming is what keeps them apart - *"these are holdings told to a
representative, never a broker account"* - written one increment before the
collision could happen. `DISPLAY_NAME` is the other half of that sentence.

## Capabilities, not a role that means "and therefore everything"

§5.3 forbids `if superuser: skip all ownership checks`, and this takes it
literally. There are two capabilities, they are declared here next to what they
gate, and they are **checked**.

What they are not is a way *around* ownership. A backend user who holds
`CAP_VIEW_SUPERUSER` and asks for the superuser portfolio gets
`OwnerContext(SUPERUSER, their own username)` - their own domain, which for
anybody but the owner is empty. The capability decides whether the question may
be asked at all; `portfolios.resolve` decides whose answer comes back. Neither
substitutes for the other, and there is no branch where holding the capability
changes the comparison.

## Nothing here is priced, and nothing is fabricated

`MANUAL` / `MANUAL` (§4.5): the holdings come from a spreadsheet somebody
maintains by hand. `is_priced()` stays LIVE-only, so the operator's portfolio is
not valued in this increment either - `data/portfolio.xlsx` carries a purchase
price and no market price, and that does not change here.

`Account ID` is **dropped rather than carried** (§9.1). `app/privacy_filter.py`
strips it on egress because it does not own that file's schema; TQ-44's holdings
table has no account column at all, which is the stronger form §96 chose
deliberately: *a field that does not exist cannot be leaked by a future reader
who forgets to sanitize.*
"""

from __future__ import annotations

from pathlib import Path

from backend import holdings, portfolio_providers, portfolios
from backend.db import Database

# --- capabilities (addendum 44 §10, spec §4.3, §11 Q3) -----------------------------
#
# Two, not §10's seven. The other five have no consumer - `PORTFOLIO_VIEW_OWN`
# duplicates the Gateway's `CAP_HOLDINGS`, `..._MANAGE_CONNECTION` is TQ-49's,
# `..._VIEW_SIMULATION` is already covered by the `simulated` flag - and the
# standing rule is that machinery with no user does not get built.
#
# They are declared here, beside what they gate, rather than in
# `gateway/roles.py` as the spec's file table guessed. That table was written
# before TQ-69 moved the subsystem; under §11 Q4 the surface is backend-side, and
# a capability declared where nothing checks it is exactly the machinery-with-no-
# user this project refuses.

CAP_VIEW_SUPERUSER = "portfolio_view_superuser"
CAP_ANALYZE_SUPERUSER = "portfolio_analyze_superuser"
CAPABILITIES = (CAP_VIEW_SUPERUSER, CAP_ANALYZE_SUPERUSER)

# Where a grant comes from: the per-user permission the operator already grants
# explicitly, through `/permissions/grant`, the CLI and the desktop.
#
# Reusing it rather than building a second grant surface, because there is
# nothing wrong with the first one: it is per-user, explicit, persisted, and
# never inferred from conversation. A parallel mechanism would be a second answer
# to "may this account reach the portfolio", and the two would drift.
#
# Both capabilities come from the one grant today, and they stay **two names
# checked separately** rather than collapsing into one. They describe different
# acts - reading a position, and having it analysed - and the second is the one
# whose output ends up in front of a model. The day something needs to grant one
# without the other, the checks already support it; what would not have supported
# it is a single capability that had to be split afterwards.
PERMISSION_RESOURCE = "portfolio"

# What the operator's own portfolio is called, so that a listing showing both of
# their portfolios is readable rather than a puzzle (§4.4).
DISPLAY_NAME = "Superuser Portfolio"

# The spreadsheet this portfolio was migrated out of. Left on disk and marked
# legacy rather than deleted (§16.1), the same habit as `client_holdings_legacy`
# and `portfolio_holdings_pre45`.
LEGACY_SPREADSHEET = Path(__file__).resolve().parent.parent / "data" / "portfolio.xlsx"

# The spreadsheet's columns, in the order it stores them, mapped to the canonical
# holding shape (TQ-45a). `Account ID` is deliberately absent - see the module
# docstring and spec §9.1.
_SHEET_COLUMNS = ("ticker", "shares", "purchase_price", "purchase_date")


class NotPermitted(PermissionError):
    """A backend user reached for the superuser portfolio without the capability.

    Its own type rather than `portfolios.NotAuthorized`, because the two answer
    different questions and only one of them has to be uninformative. §9.3's
    single refusal exists so that a caller cannot learn *whether somebody else's
    portfolio exists*; "you do not hold this capability" reveals nothing about
    anybody's data, and telling an operator plainly why a tool refused is worth
    more than a uniform silence that would send them hunting."""


def granted_for(permissions) -> frozenset[str]:
    """Which of this module's capabilities a user's explicit grants amount to.

    One place turns a permission into capabilities, so there is one answer to
    "may this account reach the Superuser Portfolio" rather than one per
    caller."""
    if permissions.is_granted(PERMISSION_RESOURCE):
        return frozenset(CAPABILITIES)
    return frozenset()


def capabilities_for(granted) -> frozenset[str]:
    """The subset of this module's capabilities a caller actually holds.

    Takes whatever the caller was granted rather than a username, so that *how*
    capabilities are granted stays somebody else's question. This module's job is
    to say which ones exist and to refuse without one."""
    return frozenset(c for c in granted if c in CAPABILITIES)


def require(granted, capability: str) -> None:
    if capability not in CAPABILITIES:
        raise ValueError(
            f"unknown superuser portfolio capability {capability!r}; known are "
            f"{list(CAPABILITIES)}")
    if capability not in capabilities_for(granted):
        raise NotPermitted(
            f"This account does not hold {capability!r}, so it cannot reach the "
            "Superuser Portfolio.")


def owner_for(username: str) -> portfolios.OwnerContext:
    """The owner context for a backend user's own superuser domain.

    **Their own**, always - there is no argument here that could name somebody
    else's, which is what makes the capability a gate on the *question* rather
    than a key to anybody's answer."""
    if not (username or "").strip():
        raise portfolios.UnknownVocabulary(
            "the Superuser Portfolio belongs to a resolved backend user; there is no "
            "owner to fall back to.")
    return portfolios.for_superuser(username)


def ensure(conn: Database, username: str) -> dict:
    """This user's superuser portfolio, created on first use.

    `MANUAL` / `MANUAL` because the holdings come from a spreadsheet a person
    maintains (§4.5), and named so that an operator holding two portfolios can
    tell which is which (§4.4).

    Through `portfolios.primary_for`, so creation goes down the same path
    everything else does and the portfolio is read back through the guard rather
    than returned as it was built."""
    return portfolios.primary_for(
        conn, owner_for(username), display_name=DISPLAY_NAME,
        provider_type=portfolios.PROVIDER_MANUAL,
        data_mode=portfolios.MODE_MANUAL)


def _read_spreadsheet(path: Path) -> list[dict]:
    """The rows, with `Account ID` never read into memory at all.

    `app/tools/portfolio.py` used to read all five columns and hand them to
    `sanitize_portfolio_rows`, which dropped the account id on the way out. That
    was right for a filter that does not own the schema. Here the destination
    schema is ours and has no account column, so the stronger form is available
    and taken: the value is not carried, not stored, and not sanitized - it is
    never picked up."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    sheet = workbook.active
    rows = [
        dict(zip(_SHEET_COLUMNS, row[:len(_SHEET_COLUMNS)]))
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row and row[0] is not None
    ]
    workbook.close()
    return rows


def migrate_spreadsheet(conn: Database, username: str,
                        path: Path | None = None) -> dict:
    """Move `data/portfolio.xlsx` into an owned portfolio, once (§16.4).

    **Idempotent through `provider_account_ref` rather than a flag.** A portfolio
    that already names its source has been migrated, so a second run finds
    nothing to do - and that is not tidiness. Re-running would re-upsert every
    spreadsheet row, overwriting edits the operator had made to those symbols and
    resurrecting the ones they had deleted.

    Nothing is valued and nothing is classified: `average_cost` is the purchase
    price per share, which is what the column already meant (§100), and
    `asset_class` is `unknown` because the sheet does not say. Guessing `stock`
    for `VTI` - an ETF - is exactly the fabrication §100's rename refused when it
    declined to map `EQUITY` to a house code."""
    source = Path(path) if path is not None else LEGACY_SPREADSHEET
    portfolio = ensure(conn, username)
    owner = owner_for(username)

    if portfolio["provider_account_ref"]:
        return {"migrated": False, "reason": "already migrated",
                "source": portfolio["provider_account_ref"], "holdings": 0}
    if not source.exists():
        return {"migrated": False, "reason": f"no spreadsheet at {source}",
                "source": None, "holdings": 0}

    rows = _read_spreadsheet(source)
    for row in rows:
        holdings.record(
            conn, portfolio,
            symbol=row["ticker"],
            quantity=row["shares"],
            average_cost=row["purchase_price"],
            # When the operator bought it is a fact about them, so it is carried
            # as `acquired_on`. `as_of` defaults to now, which is the truth about
            # this row: the data is as of the moment it was migrated, because the
            # spreadsheet does not say when it was last correct.
            acquired_on=str(row["purchase_date"]) if row["purchase_date"] else None,
        )

    # Recorded last, so an interrupted run leaves the source unrecorded and is
    # re-runnable rather than half-marked-done.
    portfolios.record_source(conn, portfolio["portfolio_id"], owner, str(source))
    return {"migrated": True, "reason": None, "source": str(source),
            "holdings": len(rows), "portfolio": portfolio["portfolio_id"]}


def holdings_for(conn: Database, username: str, granted) -> list[dict]:
    """The operator's own holdings, behind the capability and behind the guard.

    Read through the provider (TQ-45b) like every other portfolio, so the
    operator's holdings arrive by the same path a client's do rather than through
    a second reader that could disagree with the first."""
    require(granted, CAP_VIEW_SUPERUSER)
    portfolio = ensure(conn, username)
    provider = portfolio_providers.for_portfolio(portfolio)
    return [vars(h) for h in provider.get_holdings(conn, portfolio)]


def analysis_for(conn: Database, username: str, granted) -> dict:
    """Weights and concentration over the operator's own holdings.

    Its own capability rather than a widening of the view one, because §10 lists
    them separately and because reading a position and having it analysed are
    different acts - the second is the one that ends up in front of a model."""
    require(granted, CAP_ANALYZE_SUPERUSER)
    portfolio = ensure(conn, username)
    provider = portfolio_providers.for_portfolio(portfolio)
    return holdings.concentration(provider.get_holdings(conn, portfolio))
