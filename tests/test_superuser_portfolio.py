"""The operator's own portfolio, owned (TQ-46; addendum 44 §4, §10, §16, §21.4,
§21.6).

Two properties, and they fail in opposite directions — which is why both are
here.

**A client must never receive the operator's portfolio.** That is §15.5's
permanent regression and it already has tests in `test_backend_portfolios.py` and
`test_portfolio_surface.py`. It is a leak, and a leak is loud once anybody looks.

**The operator must never silently receive nobody's.** That one is new, and it is
the failure TQ-46 §11 Q4 was flagged for. With two owner domains and two possible
operator ids, a mismatch is *not refused* — `resolve()` correctly answers "not
yours", which addendum 44 §9.3 makes indistinguishable from "no such portfolio",
so the operator is shown an **empty portfolio** and a working-looking system.
`test_the_operator_is_never_silently_shown_an_empty_portfolio` is the test with
that shape, and it is the one to read first.

The owner settled which id, on 2026-08-26: the backend `username`. *"Krish is the
superuser of this system… the architect… the owner… the creator. The system
serves Krish primarily. The system also has other clients… However, they are just
clients."* And the two doors naming the operator identically is **intentional
design, not a coincidence** — one person, two doors — which is what
`test_the_gateways_operator_and_the_backends_are_one_identity` holds in place.
"""

import pytest

from backend import holdings, portfolios, superuser_portfolio
from gateway import auth as gateway_auth
from gateway import clients as gateway_clients

OPERATOR = "krish"
OTHER = "someone-else"


# --- the operator's own domain ------------------------------------------------------


def test_the_superuser_portfolio_is_owned_and_named_apart(portfolio_conn):
    """§4.4. The operator can hold two portfolios — the holdings they stated
    through the Gateway as a client, and this one — and being shown the wrong
    money is its own failure even when nothing leaked."""
    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)

    assert (mine["owner_type"], mine["owner_id"]) == (portfolios.OWNER_SUPERUSER, OPERATOR)
    assert mine["display_name"] == superuser_portfolio.DISPLAY_NAME

    # The same person's client-side portfolio, which must be a different entity.
    as_client = portfolios.primary_for(portfolio_conn, portfolios.for_client(OPERATOR))
    assert as_client["portfolio_id"] != mine["portfolio_id"]
    assert as_client["display_name"] != mine["display_name"], (
        "an operator reading a listing cannot tell their two portfolios apart")


def test_the_superuser_portfolio_is_manual_and_unpriced(portfolio_conn):
    """§4.5. The holdings come from a spreadsheet somebody maintains by hand, and
    `is_priced()` stays LIVE-only — a purchase price is not a market price."""
    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)

    assert mine["provider_type"] == portfolios.PROVIDER_MANUAL
    assert mine["data_mode"] == portfolios.MODE_MANUAL
    assert portfolios.is_priced(mine) is False


def test_ensure_is_stable_across_calls(portfolio_conn):
    first = superuser_portfolio.ensure(portfolio_conn, OPERATOR)
    second = superuser_portfolio.ensure(portfolio_conn, OPERATOR)
    assert first["portfolio_id"] == second["portfolio_id"]


def test_a_client_cannot_reach_the_superuser_portfolio(portfolio_conn):
    """§21.6, by id and by listing. `_owned_by` compares `(owner_type, owner_id)`
    as a pair, so one name in two domains is two owners — including when the name
    is the operator's own."""
    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)
    holdings.record(portfolio_conn, mine, symbol="AAPL", quantity=25, average_cost=172.34)

    for client in ("avery", OPERATOR):
        owner = portfolios.for_client(client)
        with pytest.raises(portfolios.NotAuthorized):
            portfolios.resolve(portfolio_conn, mine["portfolio_id"], owner)
        assert mine["portfolio_id"] not in {
            p["portfolio_id"] for p in portfolios.listing(portfolio_conn, owner)}


def test_one_operators_domain_does_not_reach_anothers(portfolio_conn):
    """There is no argument to `owner_for` that names somebody else, so a second
    backend user gets their own SUPERUSER domain rather than the owner's."""
    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)
    holdings.record(portfolio_conn, mine, symbol="AAPL", quantity=25, average_cost=172.34)
    theirs = superuser_portfolio.ensure(portfolio_conn, OTHER)

    assert theirs["portfolio_id"] != mine["portfolio_id"]
    with pytest.raises(portfolios.NotAuthorized):
        portfolios.resolve(portfolio_conn, mine["portfolio_id"],
                           superuser_portfolio.owner_for(OTHER))


def test_an_owner_is_required_rather_than_defaulted(portfolio_conn):
    """`for_superuser` used to default to the literal `"operator"`. A literal
    cannot be wrong in a way anybody notices, and this one would have been
    *nearly* right — which is worse."""
    import inspect

    assert inspect.signature(portfolios.for_superuser).parameters[
        "operator_id"].default is inspect.Parameter.empty

    for nobody in (None, "", "   "):
        with pytest.raises(portfolios.UnknownVocabulary):
            superuser_portfolio.owner_for(nobody)


# --- capabilities, checked rather than inferred (§5.3, §10) -------------------------


def test_the_capability_is_checked_not_inferred(portfolio_conn):
    """§5.3: `if superuser: skip all ownership checks` is not the
    implementation. Holding no capability refuses; holding one does not widen
    what may be reached."""
    superuser_portfolio.ensure(portfolio_conn, OPERATOR)

    with pytest.raises(superuser_portfolio.NotPermitted):
        superuser_portfolio.holdings_for(portfolio_conn, OPERATOR, granted=frozenset())
    with pytest.raises(superuser_portfolio.NotPermitted):
        superuser_portfolio.analysis_for(portfolio_conn, OPERATOR, granted=frozenset())


def test_view_and_analyse_are_separate_capabilities(portfolio_conn):
    """Two names, checked separately. Reading a position and having it analysed
    are different acts, and the second is the one whose output ends up in front
    of a model."""
    superuser_portfolio.ensure(portfolio_conn, OPERATOR)
    view_only = frozenset({superuser_portfolio.CAP_VIEW_SUPERUSER})

    assert superuser_portfolio.holdings_for(portfolio_conn, OPERATOR, view_only) == []
    with pytest.raises(superuser_portfolio.NotPermitted):
        superuser_portfolio.analysis_for(portfolio_conn, OPERATOR, view_only)


def test_a_capability_does_not_reach_another_owners_portfolio(portfolio_conn):
    """The capability gates the *question*; `portfolios.resolve` decides whose
    answer comes back. Neither substitutes for the other, and holding the
    capability changes no comparison."""
    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)
    holdings.record(portfolio_conn, mine, symbol="AAPL", quantity=25, average_cost=172.34)

    everything = frozenset(superuser_portfolio.CAPABILITIES)
    assert superuser_portfolio.holdings_for(portfolio_conn, OTHER, everything) == []


def test_an_unknown_capability_is_refused_rather_than_ignored(portfolio_conn):
    with pytest.raises(ValueError):
        superuser_portfolio.require(frozenset({"portfolio_do_anything"}),
                                    "portfolio_do_anything")


def test_a_grant_produces_both_capabilities_and_no_grant_produces_none(permissions_store):
    assert superuser_portfolio.granted_for(permissions_store) == frozenset()
    permissions_store.grant(superuser_portfolio.PERMISSION_RESOURCE)
    assert superuser_portfolio.granted_for(permissions_store) == frozenset(
        superuser_portfolio.CAPABILITIES)


# --- the migration (§16.4) ----------------------------------------------------------


def _sheet(tmp_path, rows):
    from openpyxl import Workbook

    path = tmp_path / "portfolio.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Ticker", "Shares", "Purchase Price", "Purchase Date", "Account ID"])
    for row in rows:
        sheet.append(list(row))
    workbook.save(path)
    return path


def test_every_spreadsheet_row_lands_with_its_own_figures(portfolio_conn, tmp_path):
    path = _sheet(tmp_path, [
        ("AAPL", 25, 172.34, "2023-03-14", "ACCT-88421"),
        ("VTI", 40, 215.67, "2022-11-09", "ACCT-88421"),
    ])

    outcome = superuser_portfolio.migrate_spreadsheet(portfolio_conn, OPERATOR, path)

    assert (outcome["migrated"], outcome["holdings"]) == (True, 2)
    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)
    held = {h["symbol"]: h for h in holdings.listing(portfolio_conn, mine)}
    assert held["AAPL"]["quantity"] == 25
    # Per share, which is what the column already meant (§100).
    assert held["AAPL"]["average_cost"] == 172.34
    assert held["AAPL"]["acquired_on"] == "2023-03-14"


def test_the_account_id_is_not_migrated_or_stored(portfolio_conn, tmp_path):
    """§9.1. Not stripped on the way out — **never carried**. A field that does
    not exist cannot be leaked by a future reader who forgets to sanitize."""
    path = _sheet(tmp_path, [("AAPL", 25, 172.34, "2023-03-14", "ACCT-88421")])

    superuser_portfolio.migrate_spreadsheet(portfolio_conn, OPERATOR, path)

    columns = {row["name"] for row in
               portfolio_conn.fetchall("PRAGMA table_info(portfolio_holdings)")}
    assert not any("account" in column for column in columns)
    dumped = str([dict(r) for r in portfolio_conn.fetchall("SELECT * FROM portfolio_holdings")])
    assert "ACCT-88421" not in dumped


def test_migrated_rows_are_not_given_an_asset_class_the_sheet_never_stated(portfolio_conn,
                                                                          tmp_path):
    """`VTI` is an ETF and `AAPL` is a stock, and the spreadsheet says neither.
    Guessing `stock` for both is the fabrication §100's rename already refused
    when it declined to map `EQUITY` to a house code."""
    path = _sheet(tmp_path, [("AAPL", 25, 172.34, "2023-03-14", "ACCT-88421"),
                             ("VTI", 40, 215.67, "2022-11-09", "ACCT-88421")])

    superuser_portfolio.migrate_spreadsheet(portfolio_conn, OPERATOR, path)

    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)
    assert {h["asset_class"] for h in holdings.listing(portfolio_conn, mine)} == {
        holdings.ASSET_UNKNOWN}


def test_the_migration_is_idempotent_and_does_not_undo_corrections(portfolio_conn, tmp_path):
    """Idempotency here is not tidiness. A second run would re-upsert every
    spreadsheet row — overwriting edits the operator had made to those symbols,
    and resurrecting the ones they had deleted."""
    path = _sheet(tmp_path, [("AAPL", 25, 172.34, "2023-03-14", "ACCT-88421")])
    superuser_portfolio.migrate_spreadsheet(portfolio_conn, OPERATOR, path)
    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)

    # The operator corrects the position, then forgets one.
    holdings.record(portfolio_conn, mine, symbol="AAPL", quantity=30, average_cost=172.34)

    again = superuser_portfolio.migrate_spreadsheet(portfolio_conn, OPERATOR, path)

    assert again["migrated"] is False
    assert again["reason"] == "already migrated"
    assert holdings.listing(portfolio_conn, mine)[0]["quantity"] == 30, (
        "a second migration undid the operator's own correction")


def test_the_migration_records_where_the_holdings_came_from(portfolio_conn, tmp_path):
    """`provider_account_ref` means "the source's own reference", so a spreadsheet
    path is exactly what belongs there — and using the field for what it means is
    what makes the migration idempotent without new machinery."""
    path = _sheet(tmp_path, [("AAPL", 25, 172.34, "2023-03-14", "ACCT-88421")])

    superuser_portfolio.migrate_spreadsheet(portfolio_conn, OPERATOR, path)

    assert superuser_portfolio.ensure(portfolio_conn, OPERATOR)["provider_account_ref"] == str(path)


def test_a_portfolio_cannot_be_repointed_at_a_different_source(portfolio_conn, tmp_path):
    """A portfolio that changed source would be a different portfolio wearing the
    same id."""
    first = _sheet(tmp_path, [("AAPL", 25, 172.34, "2023-03-14", "ACCT-1")])
    superuser_portfolio.migrate_spreadsheet(portfolio_conn, OPERATOR, first)
    mine = superuser_portfolio.ensure(portfolio_conn, OPERATOR)

    with pytest.raises(portfolios.UnknownVocabulary):
        portfolios.record_source(portfolio_conn, mine["portfolio_id"],
                                 superuser_portfolio.owner_for(OPERATOR), "somewhere/else.xlsx")


def test_a_missing_spreadsheet_is_reported_rather_than_invented(portfolio_conn, tmp_path):
    outcome = superuser_portfolio.migrate_spreadsheet(
        portfolio_conn, OPERATOR, tmp_path / "not-there.xlsx")

    assert outcome["migrated"] is False
    assert "no spreadsheet" in outcome["reason"]
    assert outcome["holdings"] == 0


# --- the failure that reads as a working system ------------------------------------


def test_the_operator_is_never_silently_shown_an_empty_portfolio(portfolio_conn, tmp_path):
    """**The test §11 Q4 asked for, and the one with the opposite shape to
    §15.5's.**

    §15.5 covers a client receiving the operator's portfolio — a leak, and loud.
    This covers the operator receiving *nobody's*, which is silent: `resolve()`
    answering "not yours" is indistinguishable from "no such portfolio" by design
    (addendum 44 §9.3), so a mismatched id produces an empty list and a
    working-looking system.

    The guard against it is that exactly one caller asserts `SUPERUSER` and it
    asserts the same id that the migration stamped. Asserted here by migrating
    under one identity and reading under it, and by showing what the *wrong* id
    would have produced — so the test fails if those two ever stop being the
    same thing."""
    path = _sheet(tmp_path, [("AAPL", 25, 172.34, "2023-03-14", "ACCT-88421"),
                             ("VTI", 40, 215.67, "2022-11-09", "ACCT-88421")])
    superuser_portfolio.migrate_spreadsheet(portfolio_conn, OPERATOR, path)
    everything = frozenset(superuser_portfolio.CAPABILITIES)

    held = superuser_portfolio.holdings_for(portfolio_conn, OPERATOR, everything)
    assert [h["symbol"] for h in held] == ["AAPL", "VTI"], (
        "the operator read their own portfolio and it came back empty - which is "
        "what a mismatched owner id looks like, because it is not an error")

    # What the wrong id would have produced: no exception, no warning, no rows.
    wrong = superuser_portfolio.holdings_for(portfolio_conn, "operator", everything)
    assert wrong == [], (
        "this assertion documents the failure mode rather than a requirement: a "
        "wrong operator id is answered with silence, so nothing but a test like "
        "this one would notice it")


# --- the ownerless retrieval is gone (§16.7, §21.1) --------------------------------


def test_no_runtime_path_reads_the_legacy_spreadsheet():
    """§8 criterion 4: `data/portfolio.xlsx` is migrated, marked legacy, and read
    by nothing.

    Exactly one function may open it — the migration, once. Anything else reading
    it would be a second source of truth for the operator's holdings, and the one
    that disagreed quietly would be the file, because nothing writes to it any
    more.

    Scanned by source rather than trusted, in the style of
    `test_nothing_outside_portfolios_queries_the_portfolios_table`. `load_workbook`
    is the only way in, so it is what the scan looks for."""
    from pathlib import Path

    from conftest import executable_source

    root = Path(__file__).resolve().parent.parent
    allowed = {("backend", "superuser_portfolio.py")}
    offenders = []
    for tree in ("app", "backend", "gateway"):
        for path in sorted((root / tree).rglob("*.py")):
            if (tree, path.name) in allowed:
                continue
            if "load_workbook" in executable_source(path):
                offenders.append(f"{tree}/{path.relative_to(root / tree)}")

    assert not offenders, (
        "these open a workbook outside the one-time migration:\n  "
        + "\n  ".join(offenders)
        + "\nThe operator's holdings live in an owned portfolio now. A second reader "
          "of data/portfolio.xlsx would be a second source of truth, and the one that "
          "went stale would be the file."
    )


def test_no_portfolio_retrieval_takes_no_owner():
    """§16.7: *"remove any global get_current_portfolio() behavior that has no
    owner argument"*.

    Every public function in `backend/superuser_portfolio.py` that reaches
    holdings names a `username`, and `retrieve_portfolio` does too. Checked from
    the signatures rather than from a source string, for the reason TQ-69's
    import tripwire had to learn: matching how somebody happened to spell
    something is guessing."""
    import inspect

    from app.tools.portfolio import retrieve_portfolio

    reaches_data = ("ensure", "holdings_for", "analysis_for", "migrate_spreadsheet",
                    "owner_for")
    for name in reaches_data:
        parameters = inspect.signature(getattr(superuser_portfolio, name)).parameters
        assert "username" in parameters, f"superuser_portfolio.{name} has no owner"
        assert parameters["username"].default is inspect.Parameter.empty, (
            f"superuser_portfolio.{name}'s owner has a default, so an ownerless call "
            "is constructible")

    parameters = inspect.signature(retrieve_portfolio).parameters
    assert "username" in parameters
    assert parameters["username"].default is inspect.Parameter.empty


# --- one identity, two doors (owner direction, 2026-08-26) --------------------------


def test_the_superuser_alias_resolves_to_the_configured_operator(monkeypatch):
    """The owner's design: a fixed `superuser` login at the Gateway, used as a
    **pseudo-id** for the operator rather than as a second account.

    The two must produce one subject. If they produced two, the operator would
    own two of everything and be shown one - the §11 Q4 failure, arriving through
    the door instead of through the migration."""
    monkeypatch.setenv(gateway_auth.SUPER_USER_ENV, OPERATOR)

    typed_own_name = gateway_auth.subject_for("operator", OPERATOR)
    typed_the_alias = gateway_auth.subject_for("operator", gateway_auth.SUPERUSER_ALIAS)

    assert typed_own_name == typed_the_alias == OPERATOR


def test_both_operator_logins_are_accepted(monkeypatch, gateway_client, gateway_conn):
    """Through the real login route, so the alias is exercised where it is
    actually used rather than only in the function that implements it."""
    from tests.conftest import GATEWAY_TEST_PASSWORD, GATEWAY_TEST_USER

    subjects = set()
    for typed in (GATEWAY_TEST_USER, gateway_auth.SUPERUSER_ALIAS):
        response = gateway_client.post(
            "/auth/login", json={"username": typed, "password": GATEWAY_TEST_PASSWORD})
        assert response.status_code == 200, f"{typed!r} could not log in"
        assert response.json()["role"] == "operator"
        subjects.add(gateway_conn.fetchone(
            "SELECT subject FROM sessions ORDER BY rowid DESC LIMIT 1")["subject"])

    assert subjects == {GATEWAY_TEST_USER}, (
        f"the alias produced a different subject: {subjects}. One person, two "
        "doors, one identity - or the operator owns two of everything.")


def test_the_alias_still_needs_the_operators_password(gateway_client):
    response = gateway_client.post(
        "/auth/login",
        json={"username": gateway_auth.SUPERUSER_ALIAS, "password": "not-the-password"})
    assert response.status_code == 401


def test_no_client_may_register_under_the_alias(gateway_conn):
    """Reserved unconditionally, not only when a credential is configured: it is
    a fixed word rather than a deployment's choice, so a client holding it on an
    unconfigured Gateway would sit waiting to collide."""
    with pytest.raises(gateway_clients.ClientRefused):
        gateway_clients.register(gateway_conn, gateway_auth.SUPERUSER_ALIAS)


def test_the_gateways_operator_and_the_backends_are_one_identity(monkeypatch):
    """**Intentional design, not a coincidence** — owner direction, 2026-08-26.

    `GATEWAY_SUPER_USER` and the backend username name the same person on
    purpose: one identity, two doors. This asserts the mechanism that makes the
    sameness *hold* rather than merely happen to be true - the Gateway records
    the configured operator name as the subject, so whatever the backend's
    `users.json` calls them is the string both sides use."""
    monkeypatch.setenv(gateway_auth.SUPER_USER_ENV, "Krish")

    subject = gateway_auth.subject_for("operator", gateway_auth.SUPERUSER_ALIAS)

    # Normalised the same way `portfolios.normalise` does, so the two sides
    # cannot disagree over case - which is the whole reason there is one
    # normalisation and not two.
    assert subject == portfolios.normalise("Krish") == OPERATOR
    assert superuser_portfolio.owner_for(subject).owner_id == OPERATOR
