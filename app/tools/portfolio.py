"""retrieve_portfolio: the one action `/chat` exposes to the model — an owned
read of an **external source**, storing nothing (TQ-46, TQ-72; §111, §112).

## What this function has survived

It began as addendum 44 §16's "legacy single-portfolio design": no owner, read
`data/portfolio.xlsx`, return whatever was in it. TQ-46 gave it an owner. TQ-72
then removed the thing TQ-46 had pointed it at — a stored `SUPERUSER` portfolio —
because the owner ruled that portfolios do not live here at all:

> §112: *"my portfolio is also external, fetched not stored."*

So the spreadsheet is not migrated into a table and read back from one. **It is a
source**, of the plainest kind: one somebody maintains by hand, whose positions
arrive with the request rather than from an API. `ManualPortfolioProvider` exists
for exactly that, and this is its first real caller.

The shape is now the shape production will have (§115) — name a source, obtain
positions from it, analyse, retain nothing — with the one simplification that a
locally-maintained file needs no credentials. When TQ-73 builds the credentialed
fetch, this path changes where the positions come from and nothing else.

## Two layers, and the second one is the whole guarantee

  Layer 1 (`permissions.py`) — may My AI reach this account's portfolio at all.
  Layer 2 (`privacy_preferences.py`) — may those positions be forwarded to the
    reasoning model. With no stored disposition this returns `needs_consent`
    rather than data or a denial, and `/chat` pauses and asks.

TQ-46 §11 Q2 asked whether layer 2 was redundant now that §108 refuses a
`LOCAL_ONLY` task rather than forwarding it. It is not, and the check inverted
the expectation: `PORTFOLIO_FIELD_CLASSES` marks only `account_id` as
`LOCAL_ONLY`, and that field is **never read** here — so nothing in what reaches
a model carries a `LOCAL_ONLY` classification, `privacy_floor_for()` returns
`None`, and `PATH_REFUSED` never fires.

**This prompt is the only control on the operator's positions leaving the
machine.** Removing it as duplicated would have left nothing, which is why TQ-46
kept it and why TQ-72 did not quietly take it away while removing everything
around it.

## The account id is never read, not stripped

`app/privacy_filter.py` strips `account_id` on egress because it does not own
that file's schema. Here the reader owns what it picks up, so the stronger form
is available and taken: the column is **not carried into memory at all**. A value
that was never read cannot be leaked by a future caller who forgets to sanitize.
"""

from pathlib import Path

from ..audit import AuditLog
from ..permissions import PermissionManager
from ..privacy_preferences import PrivacyPreferenceStore
from ..users import ensure_user_data_dir

# What an account's own portfolio source is called, inside its own directory.
#
# **Per user, derived from the authenticated username, and never an argument.**
# This is the sharpest thing in the file. The old version read one path -
# `RESOURCE_PATHS["portfolio"]`, the same file for everybody - and
# `tests/test_multi_user_isolation.py` recorded that as deliberate: *"portfolio.xlsx
# stays a single shared file by design - only governance state is per-user."*
#
# It was not a design, it was the ownerless retrieval wearing a second hat. A
# fully-granted second account read the first account's positions, and nothing
# was wrong with the code. TQ-46 gave the *function* an owner; this gives the
# *source* one, which is where it actually mattered.
#
# There is no fallback to a shared location, and that is the point. A fallback is
# how one account ends up reading another's file the day their own is missing -
# and "missing" would be the moment it happened, which is the least likely moment
# for anybody to be watching.
SOURCE_FILENAME = "portfolio.xlsx"

RESOURCE = "portfolio"
FORWARDING_KEY = "portfolio_holdings:reasoning_model"
CONSENT_PROMPT = (
    "I need to share your portfolio positions (symbol, quantity, average cost, "
    "acquisition date) with the reasoning model to answer this. Allow this?"
)

# The spreadsheet's columns, in the order it stores them, mapped to the canonical
# holding shape. `Account ID` is deliberately absent - see the module docstring.
_SHEET_COLUMNS = ("symbol", "quantity", "average_cost", "acquired_on")


def _read_source(path: Path):
    """The operator's own source, read at request time and held nowhere.

    Returns a `portfolio_providers.Source` carrying the positions, because a
    hand-maintained file *is* its positions - there is no API behind it to fetch
    from later."""
    from openpyxl import load_workbook

    from backend import portfolio_providers, portfolios

    workbook = load_workbook(path, read_only=True)
    sheet = workbook.active
    positions = tuple(
        dict(zip(_SHEET_COLUMNS, row[:len(_SHEET_COLUMNS)]))
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row and row[0] is not None
    )
    workbook.close()

    return portfolio_providers.Source(
        provider_type=portfolios.PROVIDER_MANUAL,
        name=path.name,
        reference=str(path),
        data_mode=portfolios.MODE_MANUAL,
        simulated=False,
        positions=positions,
    )


def retrieve_portfolio(
    conn,
    username: str,
    permissions: PermissionManager,
    preferences: PrivacyPreferenceStore,
    audit_log: AuditLog,
    allow_once: bool = False,
) -> dict:
    """This account's own portfolio positions, from its own external source.

    `username` is required and comes from `Depends(get_current_user)` — the value
    `/chat` was already resolving and handing nowhere near this function (§110
    §3). §16.7's "ownerless retrieval" was never a missing identity mechanism; it
    was one being discarded.

    There is no argument here that could name somebody else's portfolio, and
    since §111 there is no stored portfolio for one to name. The owner is the
    caller, always.

    `conn` is unused and kept: TQ-73's fetch will need it for the market data
    store (§113), and a signature that loses it now would have to grow it back.
    Recorded rather than silently ignored, because an unused argument nobody
    explains is one somebody eventually removes."""
    from backend import portfolio_providers

    if not permissions.is_granted(RESOURCE):
        audit_log.record(
            action="retrieve_portfolio",
            resource=RESOURCE,
            authorized=False,
            result="denied - permission not granted",
        )
        return {"error": "Permission to access the portfolio has been revoked or was never granted."}

    disposition = preferences.get(FORWARDING_KEY)
    if disposition is None and not allow_once:
        return {"status": "needs_consent", "consent_key": FORWARDING_KEY, "prompt": CONSENT_PROMPT}

    if disposition == "never":
        audit_log.record(
            action="retrieve_portfolio",
            resource=RESOURCE,
            authorized=False,
            result="denied - forwarding disposition is 'never'",
        )
        return {"error": "You've asked me not to share portfolio data with the reasoning model."}

    path = Path(ensure_user_data_dir(username)) / SOURCE_FILENAME
    if not path.exists():
        # Absent is absent, never an empty portfolio. An empty list here would
        # read as "you hold nothing", which is a false statement about somebody's
        # money in the voice of a true one - the §110 §4.5 rule, one domain over.
        audit_log.record(
            action="retrieve_portfolio",
            resource=RESOURCE,
            authorized=True,
            result=f"source unavailable - {path}",
        )
        return {"error": "There is no portfolio source configured for this account. I am "
                         "not showing you an empty portfolio instead, and I will not "
                         "read anybody else's."}

    source = _read_source(path)
    provider = portfolio_providers.for_source(source)
    try:
        held = provider.get_holdings(source)
    except portfolio_providers.ProviderRefused as refusal:
        return {"error": str(refusal)}

    audit_log.record(
        action="retrieve_portfolio",
        resource=RESOURCE,
        authorized=True,
        result=(f"returned {len(held)} holdings"
                + (" (one-time consent, not persisted)" if disposition is None else "")),
    )
    return {"holdings": [vars(h) for h in held]}
